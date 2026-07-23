#!/usr/bin/env python3
"""Build vehicle_catalog.json (+ prompts) from Vehicles-data/India-Vehicles-Master-List.xlsx.

Uses Vehicle Class Cars + Luxury only (same scope as Vehiclegen UI).
Generation rows for the same Make+Model are collapsed into one model entry
with years from Gen Start Year values.
"""

from __future__ import annotations

import json
import re
from collections import defaultdict
from pathlib import Path

try:
    import openpyxl
except ImportError as exc:  # pragma: no cover
    raise SystemExit("openpyxl is required: pip install openpyxl") from exc

ROOT = Path(__file__).resolve().parent
XLSX = ROOT / "Vehicles-data" / "India-Vehicles-Master-List.xlsx"
CATALOG_OUT = ROOT / "vehicle_catalog.json"
PROMPTS_JSON_OUT = ROOT / "vehicle_prompts.json"
PROMPTS_TXT_OUT = ROOT / "vehicle_prompts.txt"

INCLUDED_CLASSES = {"Cars": "normal", "Luxury": "luxury"}

# New master list tags luxury two-wheelers as Vehicle Class=Luxury (no Cars subclass).
# Vehiclegen is a car list — keep the historical Cars+Luxury-cars scope.
LUXURY_TWO_WHEELER_MAKES = {
    "Ducati",
    "Harley-Davidson",
    "Indian Motorcycle",
    "MV Agusta",
    "Moto Morini",
    "Triumph",
    "Victory Motorcycles",
}

# Legacy prompt template kept for vehicle_prompts.* dumps (UI builds prompts live).
PROMPT_TEMPLATE = (
    "A photorealistic studio image of a {make} {model} in {color}, shot from the "
    "driver-side front 3/4 angle, matching the exact camera angle, height, and "
    "perspective of the @[0a99b260-c95c-4535-9af7-b8e4e90b8825:Honda Car Reference:output] . "
    "The car's wheel/tyre angle must remain perfectly straight — do not turn, angle, "
    "or steer the front wheels; match the reference image's tyre orientation strictly "
    "exactly. Lighting must replicate the "
    "@[0a99b260-c95c-4535-9af7-b8e4e90b8825:Honda Car Reference:output] : same direction, "
    "softness, and intensity of light and highlights on the body panels. Include a "
    "realistic, soft-edged contact shadow directly beneath the car where the tyres "
    "meet the ground, consistent with the lighting direction in the "
    "@[0a99b260-c95c-4535-9af7-b8e4e90b8825:Honda Car Reference:output] strictly — no "
    "floating car, no hard black blob shadow. Remove the number plate (front and rear) "
    "and leave that area clean/blank as if no plate was ever fitted. Keep all badges, "
    "logos, and emblems (grille logo, wheel center caps, model badging) . The background "
    "must be completely transparent — true alpha transparency, no white, no gray, no "
    "checkerboard pattern baked into the image, no floor, no wall, no gradient, nothing "
    "except the car and its ground shadow. Ultra-realistic automotive photography style, "
    "sharp focus, accurate reflections on paint and glass, correct tyre and rim detail, "
    "natural paint texture. keep all the details very strictly and prompt"
)


def _split_list(value: object) -> list[str]:
    if value is None:
        return []
    text = str(value).strip()
    if not text or text == "—":
        return []
    parts = []
    for raw in text.split(";"):
        item = raw.strip()
        if not item or item == "—":
            continue
        # Drop trailing research notes in parentheses when they are clearly asides.
        item = re.sub(r"\s*\([^)]*(?:also listed|dealers|approx)[^)]*\)\s*$", "", item, flags=re.I)
        item = item.strip().rstrip(",")
        if item:
            parts.append(item)
    # Preserve order, de-dupe case-insensitively
    seen: set[str] = set()
    out: list[str] = []
    for p in parts:
        key = p.casefold()
        if key in seen:
            continue
        seen.add(key)
        out.append(p)
    return out


def _year(value: object) -> int | None:
    if value is None or value == "":
        return None
    try:
        return int(value)
    except (TypeError, ValueError):
        try:
            return int(float(value))
        except (TypeError, ValueError):
            return None


def _status_rank(status: object) -> int:
    s = (str(status or "")).strip().casefold()
    order = {
        "current": 0,
        "superseded": 1,
        "discontinued": 2,
        "unknown": 3,
        "never sold new in india": 4,
    }
    return order.get(s, 9)


def load_workbook_rows():
    wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
    try:
        ws = wb["Vehicles"]
        rows = list(ws.iter_rows(values_only=True))
        header = list(rows[0])
        idx = {name: i for i, name in enumerate(header)}
        required = [
            "Model Key",
            "Vehicle Class",
            "Make",
            "Model",
            "Generation",
            "Gen Start Year",
            "Gen End Year",
            "Lifecycle Status",
            "Body / Type",
            "Official Colour Names (India)",
        ]
        missing = [c for c in required if c not in idx]
        if missing:
            raise SystemExit(f"Vehicles sheet missing columns: {missing}")

        families: dict[str, str] = {}
        if "Colour Families Archive" in wb.sheetnames:
            for r in wb["Colour Families Archive"].iter_rows(values_only=True):
                if not r or not r[0] or r[0] == "Model Key":
                    continue
                families[str(r[0]).strip()] = r[1] if len(r) > 1 else None

        data_rows = []
        for r in rows[1:]:
            if not r:
                continue
            make = r[idx["Make"]]
            model = r[idx["Model"]]
            vclass = r[idx["Vehicle Class"]]
            if not make or not model or vclass not in INCLUDED_CLASSES:
                continue
            if str(make).strip() in LUXURY_TWO_WHEELER_MAKES:
                continue
            data_rows.append(r)
        return idx, data_rows, families
    finally:
        wb.close()


def build_catalog():
    idx, data_rows, families = load_workbook_rows()
    groups: dict[tuple[str, str, str], list] = defaultdict(list)

    for r in data_rows:
        make = str(r[idx["Make"]]).strip()
        model = str(r[idx["Model"]]).strip()
        category = INCLUDED_CLASSES[r[idx["Vehicle Class"]]]
        groups[(make, model, category)].append(r)

    makes_map: dict[tuple[str, str], dict] = {}

    for (make, model, category), gens in groups.items():
        gens_sorted = sorted(
            gens,
            key=lambda r: (
                _status_rank(r[idx["Lifecycle Status"]]),
                -(_year(r[idx["Gen Start Year"]]) or 0),
                str(r[idx["Generation"]] or ""),
            ),
        )
        preferred = gens_sorted[0]

        colors: list[str] = []
        for r in gens_sorted:
            colors = _split_list(r[idx["Official Colour Names (India)"]])
            if colors:
                break
        if not colors:
            # Fall back to archived colour families so the UI still has selectable options.
            for r in gens_sorted:
                key = r[idx["Model Key"]]
                colors = _split_list(families.get(str(key).strip() if key else ""))
                if colors:
                    break

        colour_families: list[str] = []
        for r in gens_sorted:
            key = r[idx["Model Key"]]
            colour_families = _split_list(families.get(str(key).strip() if key else ""))
            if colour_families:
                break

        years = sorted(
            {
                y
                for r in gens
                if (y := _year(r[idx["Gen Start Year"]])) is not None
            }
        )
        launch_year = years[0] if years else None
        update_year = years[-1] if len(years) > 1 else None

        model_entry = {
            "name": model,
            "colors": colors,
            "colour_families": colour_families,
            "body": (str(preferred[idx["Body / Type"]]).strip() if preferred[idx["Body / Type"]] else "—"),
            "segment": "—",  # Segment column removed in master list v6
            "launch_year": launch_year,
            "update_year": update_year,
            "years": years,
            "lifecycle_status": (
                str(preferred[idx["Lifecycle Status"]]).strip()
                if preferred[idx["Lifecycle Status"]]
                else None
            ),
            "model_keys": [
                str(r[idx["Model Key"]]).strip()
                for r in sorted(
                    gens,
                    key=lambda r: (_year(r[idx["Gen Start Year"]]) or 0, str(r[idx["Generation"]] or "")),
                )
                if r[idx["Model Key"]]
            ],
        }

        make_key = (make, category)
        if make_key not in makes_map:
            makes_map[make_key] = {"name": make, "category": category, "models": []}
        makes_map[make_key]["models"].append(model_entry)

    makes = []
    for key in sorted(makes_map.keys(), key=lambda k: (k[0].casefold(), k[1])):
        entry = makes_map[key]
        entry["models"] = sorted(entry["models"], key=lambda m: m["name"].casefold())
        makes.append(entry)

    normal_makes = [m for m in makes if m["category"] == "normal"]
    luxury_makes = [m for m in makes if m["category"] == "luxury"]
    model_count = sum(len(m["models"]) for m in makes)
    models_with_year = sum(
        1 for m in makes for model in m["models"] if model.get("years")
    )

    catalog = {
        "source": "Vehicles-data/India-Vehicles-Master-List.xlsx",
        "sheets": ["Vehicles"],
        "categories": {
            "normal": "Normal (Vehicle Class = Cars)",
            "luxury": "Luxury (Vehicle Class = Luxury)",
        },
        "make_count": len(makes),
        "model_count": model_count,
        "normal_make_count": len(normal_makes),
        "luxury_make_count": len(luxury_makes),
        "models_with_year": models_with_year,
        "makes": makes,
    }
    return catalog


def build_prompts(catalog: dict) -> tuple[dict, str]:
    cars = []
    lines = [
        f"Vehicle Image Gen prompts — {catalog['model_count']} models (first color each)",
        f"Source: {catalog['source']}",
        "=" * 72,
        "",
    ]
    n = 0
    for make in catalog["makes"]:
        for model in make["models"]:
            colors = model.get("colors") or []
            if not colors:
                continue
            n += 1
            color = colors[0]
            prompt = PROMPT_TEMPLATE.format(
                make=make["name"], model=model["name"], color=color
            )
            cars.append(
                {
                    "id": n,
                    "make": make["name"],
                    "model": model["name"],
                    "color": color,
                    "colors": colors,
                    "prompt": prompt,
                }
            )
            lines.append(f"{n}. {make['name']} {model['name']} — {color}")
            lines.append(f"   Available colors: {'; '.join(colors)}")
            lines.append(prompt)
            lines.append("")
            lines.append("-" * 72)
            lines.append("")

    prompts = {
        "source": catalog["source"],
        "note": "One prompt per model using the first catalog color. UI supports all colors via Make → Model → Color.",
        "count": len(cars),
        "make_count": catalog["make_count"],
        "model_count": catalog["model_count"],
        "cars": cars,
    }
    return prompts, "\n".join(lines).rstrip() + "\n"


def main() -> None:
    if not XLSX.exists():
        raise SystemExit(f"Missing master sheet: {XLSX}")

    catalog = build_catalog()
    prompts, prompts_txt = build_prompts(catalog)

    CATALOG_OUT.write_text(json.dumps(catalog, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
    PROMPTS_JSON_OUT.write_text(json.dumps(prompts, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
    PROMPTS_TXT_OUT.write_text(prompts_txt, encoding="utf-8")

    print(
        f"Wrote {CATALOG_OUT.name}: {catalog['make_count']} makes "
        f"({catalog['luxury_make_count']} luxury · {catalog['normal_make_count']} normal) · "
        f"{catalog['model_count']} models · {catalog['models_with_year']} with years"
    )
    print(f"Wrote {PROMPTS_JSON_OUT.name} / {PROMPTS_TXT_OUT.name}: {prompts['count']} prompts")


if __name__ == "__main__":
    main()
